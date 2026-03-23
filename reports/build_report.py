import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Styles ──
gold = Font(name='Arial', bold=True, color='D4AF37', size=14)
header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='1A1A1A', end_color='1A1A1A', fill_type='solid')
data_font = Font(name='Arial', size=10)
qb_fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
rb_fill = PatternFill(start_color='E6FFE6', end_color='E6FFE6', fill_type='solid')
wr_fill = PatternFill(start_color='E6E6FF', end_color='E6E6FF', fill_type='solid')
te_fill = PatternFill(start_color='FFF0E6', end_color='FFF0E6', fill_type='solid')
idp_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
k_fill = PatternFill(start_color='FFFDE6', end_color='FFFDE6', fill_type='solid')
pos_fills = {'QB':qb_fill,'RB':rb_fill,'WR':wr_fill,'TE':te_fill,'DL':idp_fill,'LB':idp_fill,'DB':idp_fill,'K':k_fill}
thin_border = Border(
    left=Side(style='thin', color='DDDDDD'), right=Side(style='thin', color='DDDDDD'),
    top=Side(style='thin', color='DDDDDD'), bottom=Side(style='thin', color='DDDDDD')
)

# ── Read data ──
with open('/Users/jacobc/Downloads/dhq_data.txt', 'r') as f:
    content = f.read()

parts = content.split('=== PICKS ===')
player_lines = parts[0].replace('=== PLAYERS ===\n', '').strip().split('\n')
pick_lines = parts[1].strip().split('\n')

print(f"Players: {len(player_lines)-1}, Picks: {len(pick_lines)-1}")

# ══════════════════════════════════════════════════════
# SHEET 1: All Players
# ══════════════════════════════════════════════════════
ws = wb.active
ws.title = 'All Players'

# Title row
ws.merge_cells('A1:M1')
ws['A1'] = 'DHQ Engine Report — The Psycho League: Year VI (16-team SF IDP Half-PPR)'
ws['A1'].font = gold
ws['A1'].alignment = Alignment(horizontal='center')
ws.row_dimensions[1].height = 30

# Headers
headers = ['Rank','Player','Pos','Team','Age','DHQ Value','wPPG','Age Factor','Sit Mult','Peak Yrs Left','Starter Seasons','Recent GP','Trend %']
for col, h in enumerate(headers, 1):
    c = ws.cell(row=3, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal='center')
    c.border = thin_border

# Data
for i, line in enumerate(player_lines[1:], 4):  # skip header
    parts = line.split('|')
    if len(parts) < 13: continue
    vals = [int(parts[0]), parts[1], parts[2], parts[3], int(parts[4]),
            int(parts[5]), float(parts[6]), float(parts[7]), float(parts[8]),
            int(parts[9]), int(parts[10]), int(parts[11]), int(parts[12])]
    for col, v in enumerate(vals, 1):
        c = ws.cell(row=i, column=col, value=v)
        c.font = data_font
        c.border = thin_border
        if col == 3:  # pos column
            fill = pos_fills.get(v)
            if fill:
                for cc in range(1, 14):
                    ws.cell(row=i, column=cc).fill = fill

# Column widths
widths = [6, 25, 5, 5, 5, 10, 7, 10, 9, 12, 14, 10, 8]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# Freeze panes
ws.freeze_panes = 'A4'

# Number formats
for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=6, max_col=6):
    for cell in row:
        cell.number_format = '#,##0'
for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=7, max_col=7):
    for cell in row:
        cell.number_format = '0.0'
for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=8, max_col=9):
    for cell in row:
        cell.number_format = '0.000'

print(f"Players sheet: {ws.max_row - 3} rows")

# ══════════════════════════════════════════════════════
# SHEET 2: Draft Pick Values
# ══════════════════════════════════════════════════════
ws2 = wb.create_sheet('Draft Pick Values')

ws2.merge_cells('A1:H1')
ws2['A1'] = 'DHQ Draft Pick Values — Based on League Draft History (4 seasons)'
ws2['A1'].font = gold
ws2['A1'].alignment = Alignment(horizontal='center')
ws2.row_dimensions[1].height = 30

pick_headers = ['Slot','Round','Pick In Round','DHQ Value','Hit Rate %','Starter Rate %','Avg Norm PPG','Samples']
for col, h in enumerate(pick_headers, 1):
    c = ws2.cell(row=3, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal='center')
    c.border = thin_border

round_fills = {
    1: PatternFill(start_color='FFF8E1', end_color='FFF8E1', fill_type='solid'),
    2: PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid'),
    3: PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid'),
    4: PatternFill(start_color='F3E5F5', end_color='F3E5F5', fill_type='solid'),
    5: PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid'),
    6: PatternFill(start_color='ECEFF1', end_color='ECEFF1', fill_type='solid'),
    7: PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid'),
}

for i, line in enumerate(pick_lines[1:], 4):
    parts = line.split('|')
    if len(parts) < 8: continue
    vals = [int(parts[0]), int(parts[1]), int(parts[2]), int(parts[3]),
            int(parts[4]), int(parts[5]), float(parts[6]), int(parts[7])]
    rd = vals[1]
    for col, v in enumerate(vals, 1):
        c = ws2.cell(row=i, column=col, value=v)
        c.font = data_font
        c.border = thin_border
        if rd in round_fills:
            c.fill = round_fills[rd]

pick_widths = [6, 7, 13, 11, 11, 13, 13, 9]
for i, w in enumerate(pick_widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = 'A4'

for row in ws2.iter_rows(min_row=4, max_row=ws2.max_row, min_col=4, max_col=4):
    for cell in row:
        cell.number_format = '#,##0'

print(f"Picks sheet: {ws2.max_row - 3} rows")

# ══════════════════════════════════════════════════════
# SHEET 3: Formula Reference
# ══════════════════════════════════════════════════════
ws3 = wb.create_sheet('Formula Reference')
ws3.merge_cells('A1:C1')
ws3['A1'] = 'DHQ Engine Formula Breakdown'
ws3['A1'].font = gold
ws3.column_dimensions['A'].width = 25
ws3.column_dimensions['B'].width = 15
ws3.column_dimensions['C'].width = 70

ref_data = [
    ['Component', 'Weight', 'Description'],
    ['Core Score', '75%', 'wPPG × AgeFactor × SitMult, normalized to 0-7500 scale'],
    ['wPPG', '—', 'Weighted avg PPG across 5 seasons. Recent years weighted more. Best season gets bonus weight.'],
    ['AgeFactor', '—', '1.0 during peak years. Decays after position peak: QB=34, RB=27, WR=30, TE=30. Rate: 6%/year.'],
    ['SitMult', '—', 'Positional rank multiplier: Top3=1.20x, Top5=1.12x, Top10=1.05x, Bottom25%=0.88x. Clamped 0.40-1.60.'],
    ['Scarcity', '10%', 'Position premium (0-1000). QB tiered in SF: Top12=750, QB13-24=400, QB25+=100. Zero for unrostered.'],
    ['Peak Bonus', '5%', '120 DHQ per peak year remaining, capped at 1000.'],
    ['Consistency', '—', '4+ starter seasons = 400, 3 = 300, 2 = 150. Zero for unrostered players.'],
    ['Durability', '—', '16+ recent GP = 100, 13+ = 50.'],
    ['', '', ''],
    ['PICK VALUES', '', 'Based on YOUR leagues actual 4-season draft history'],
    ['Hit Rate', '', '% of picks at that slot that produced a top-15% player at their position'],
    ['Starter Rate', '', '% of picks at that slot that produced a starter-level season'],
    ['AvgNormPPG', '', 'Average normalized PPG produced by picks at that draft slot'],
    ['', '', ''],
    ['SCALE', '', 'All DHQ values are 0-10,000. Top player ~9,500. Average starter ~3,000-4,000.'],
    ['LEAGUE', '', 'The Psycho League: Year VI — 16-team Superflex IDP Half-PPR'],
    ['ENGINE', '', 'Values recalculated using YOUR league scoring settings, not generic rankings'],
]

for i, row in enumerate(ref_data, 3):
    for j, val in enumerate(row):
        c = ws3.cell(row=i, column=j+1, value=val)
        c.font = Font(name='Arial', bold=(i==3), size=11 if i==3 else 10)
        if i == 3:
            c.fill = header_fill
            c.font = header_font
        c.border = thin_border

# ── Save ──
outpath = '/Users/jacobc/Projects/reconai/reports/DHQ_Engine_Report.xlsx'
wb.save(outpath)
print(f'Saved to {outpath}')
