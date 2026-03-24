import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.load_workbook('/Users/jacobc/Projects/reconai/reports/DHQ_Engine_Report.xlsx')

# Delete old Formula Reference sheet if exists
if 'Formula Reference' in wb.sheetnames:
    del wb['Formula Reference']

ws = wb.create_sheet('Formula Reference')

# Styles
title_font = Font(name='Arial', bold=True, color='D4AF37', size=16)
section_font = Font(name='Arial', bold=True, color='D4AF37', size=13)
header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='1A1A1A', end_color='1A1A1A', fill_type='solid')
bold_font = Font(name='Arial', bold=True, size=11)
normal_font = Font(name='Arial', size=11)
small_font = Font(name='Arial', size=10, color='666666')
accent_font = Font(name='Arial', bold=True, size=11, color='2E86C1')
green_font = Font(name='Arial', bold=True, size=11, color='27AE60')
red_font = Font(name='Arial', bold=True, size=11, color='E74C3C')
gold_fill = PatternFill(start_color='FFF8E1', end_color='FFF8E1', fill_type='solid')
gray_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
thin = Border(
    left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC')
)

# Column widths
ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 28
ws.column_dimensions['C'].width = 14
ws.column_dimensions['D'].width = 65

r = 1
# ── TITLE ──
ws.merge_cells('B1:D1')
ws['B1'] = 'DHQ ENGINE — FORMULA REFERENCE'
ws['B1'].font = title_font
ws.row_dimensions[1].height = 35
r = 3

# ── OVERVIEW ──
ws.merge_cells(f'B{r}:D{r}')
ws[f'B{r}'] = 'OVERVIEW'
ws[f'B{r}'].font = section_font
ws[f'B{r}'].fill = gold_fill
r += 1
overview = [
    'The DHQ (Dynasty Headquarters) Engine calculates a value for every NFL player on a 0-10,000 scale.',
    'Unlike generic rankings, DHQ values are calculated using YOUR leagues actual scoring settings,',
    'draft history, roster construction, and positional scarcity. Every league gets different values.',
    '',
    'League: The Psycho League: Year VI | 16-team | Superflex | IDP | Half-PPR',
    'Data Sources: 5 seasons of stats, 4 seasons of draft picks, 3 seasons of FAAB transactions',
    'Total Players Scored: 1,977 | Draft Pick Slots Valued: 112 (7 rounds × 16 teams)',
]
for line in overview:
    ws[f'B{r}'] = line
    ws[f'B{r}'].font = normal_font
    ws.merge_cells(f'B{r}:D{r}')
    r += 1
r += 1

# ── MASTER FORMULA ──
ws.merge_cells(f'B{r}:D{r}')
ws[f'B{r}'] = 'MASTER FORMULA'
ws[f'B{r}'].font = section_font
ws[f'B{r}'].fill = gold_fill
r += 1
ws.merge_cells(f'B{r}:D{r}')
ws[f'B{r}'] = 'DHQ Value = Core Score + Scarcity + Peak Bonus + Consistency + Durability'
ws[f'B{r}'].font = Font(name='Courier New', bold=True, size=13, color='1A1A1A')
ws[f'B{r}'].fill = gray_fill
ws.row_dimensions[r].height = 28
r += 1
ws.merge_cells(f'B{r}:D{r}')
ws[f'B{r}'] = 'Where Core Score = (wPPG × AgeFactor × SitMult) / TopComposite × 7500'
ws[f'B{r}'].font = Font(name='Courier New', size=11, color='444444')
r += 2
# ── COMPONENT BREAKDOWN ──
ws.merge_cells(f'B{r}:D{r}')
ws[f'B{r}'] = 'COMPONENT BREAKDOWN'
ws[f'B{r}'].font = section_font
ws[f'B{r}'].fill = gold_fill
r += 1

# Headers
for col, val in [('B','Component'), ('C','Weight'), ('D','How It Works')]:
    c = ws[f'{col}{r}']
    c.value = val; c.font = header_font; c.fill = header_fill; c.border = thin
r += 1

components = [
    ['CORE SCORE', '75%', 'The primary value driver. Combines production (wPPG), age trajectory (AgeFactor), and league situation (SitMult). Normalized against the top player to a 0-7,500 scale.'],
    ['', '', ''],
    ['  wPPG (Weighted PPG)', '—', 'Weighted average of PPG across last 5 seasons. Recent seasons weighted heavier than older ones. Best single season gets extra weight to protect elite ceilings.'],
    ['  AgeFactor', '—', 'Multiplier based on age vs position-specific peak window. 1.0 during peak years. Decays at 6% per year after peak. Peak windows: QB=34, RB=27, WR=30, TE=30, IDP=30.'],
    ['  SitMult (Situation)', '—', 'League rank multiplier. Top 3 at position = 1.20×. Top 5 = 1.12×. Top 10 = 1.05×. Bottom 25% = 0.88×. Bottom 10% = 0.78×. Clamped between 0.40 and 1.60.'],
    ['', '', ''],
    ['SCARCITY', '10%', 'Positional premium on a 0-1,000 scale. In Superflex leagues, QB scarcity is tiered: Top 12 QBs = 750, QB13-24 = 400, QB25+ = 100. Other positions get a flat scarcity based on starter pool size. Unrostered players get 0.'],
    ['', '', ''],
    ['PEAK BONUS', '5%', '120 DHQ per remaining peak year, capped at 1,000. A 23-year-old WR with 7 peak years left gets 840 bonus. A 31-year-old RB past peak gets 0.'],
    ['', '', ''],
    ['CONSISTENCY', '~4%', 'Rewards proven producers. 4+ starter seasons = 400. 3 seasons = 300. 2 seasons = 150. Only applies to rostered players — unrostered get 0.'],
    ['', '', ''],
    ['DURABILITY', '~1%', 'Small bonus for availability. 16+ games played in most recent season = 100. 13+ games = 50. Rewards iron men, penalizes injury-prone.'],
]

for comp in components:
    if comp[0] == '':
        r += 1
        continue
    is_sub = comp[0].startswith('  ')
    ws[f'B{r}'] = comp[0]
    ws[f'B{r}'].font = bold_font if not is_sub else normal_font
    ws[f'C{r}'] = comp[1]
    ws[f'C{r}'].font = accent_font
    ws[f'C{r}'].alignment = Alignment(horizontal='center')
    ws[f'D{r}'] = comp[2]
    ws[f'D{r}'].font = normal_font
    ws[f'D{r}'].alignment = Alignment(wrap_text=True)
    for col in ['B','C','D']:
        ws[f'{col}{r}'].border = thin
    if not is_sub and comp[0] != '':
        for col in ['B','C','D']:
            ws[f'{col}{r}'].fill = gray_fill
    ws.row_dimensions[r].height = 45 if len(comp[2]) > 80 else 20
    r += 1
r += 1
# ── WORKED EXAMPLE ──
ws.merge_cells(f'B{r}:D{r}')
ws[f'B{r}'] = 'WORKED EXAMPLE: Josh Allen (DHQ 9,350)'
ws[f'B{r}'].font = section_font
ws[f'B{r}'].fill = gold_fill
r += 1

example = [
    ['Input', 'Value', 'Explanation'],
    ['wPPG', '25.6', 'Weighted PPG across 5 seasons — elite production'],
    ['AgeFactor', '1.000', 'Age 29, QB peak window extends to 34 — still in prime'],
    ['SitMult', '1.558', 'Top 3 QB in this leagues scoring — elite situation bonus'],
    ['Core Score', '~7,500', '(25.6 × 1.0 × 1.558) / topComposite × 7500 = ~7,500'],
    ['Scarcity', '+750', 'Top 12 QB in Superflex — full QB scarcity premium'],
    ['Peak Bonus', '+600', '5 peak years remaining × 120 = 600'],
    ['Consistency', '+400', '4 starter seasons — maximum consistency bonus'],
    ['Durability', '+100', '17 games played — full durability bonus'],
    ['TOTAL', '9,350', '7,500 + 750 + 600 + 400 + 100 = 9,350'],
]

for i, row in enumerate(example):
    for j, val in enumerate(row):
        c = ws.cell(row=r, column=j+2, value=val)
        c.border = thin
        if i == 0:
            c.font = header_font; c.fill = header_fill
        elif i == len(example)-1:
            c.font = Font(name='Arial', bold=True, size=11, color='D4AF37')
            c.fill = gold_fill
        else:
            c.font = normal_font
    r += 1
r += 1
# ── AGE CURVES ──
ws.merge_cells(f'B{r}:D{r}')
ws[f'B{r}'] = 'POSITION PEAK WINDOWS & AGE DECAY'
ws[f'B{r}'].font = section_font
ws[f'B{r}'].fill = gold_fill
r += 1

for col, val in [('B','Position'), ('C','Peak Until'), ('D','Notes')]:
    c = ws[f'{col}{r}']
    c.value = val; c.font = header_font; c.fill = header_fill; c.border = thin
r += 1

age_data = [
    ['QB', 'Age 34', 'Longest peak window. QBs hold value deep into their 30s. 6% annual decay after 34.'],
    ['RB', 'Age 27', 'Shortest peak. RBs decline fastest. A 28-year-old RB is already decaying at 6%/year.'],
    ['WR', 'Age 30', 'Middle ground. WRs can produce into early 30s but decline accelerates after 30.'],
    ['TE', 'Age 30', 'Similar to WR. Late bloomers common at TE so the engine accounts for this.'],
    ['DL/LB/DB', 'Age 30', 'IDP players follow a similar curve to offensive skill positions.'],
    ['K', 'Age 30', 'Kickers have long careers but limited dynasty value regardless of age.'],
]

for row in age_data:
    ws[f'B{r}'] = row[0]; ws[f'B{r}'].font = bold_font; ws[f'B{r}'].border = thin
    ws[f'C{r}'] = row[1]; ws[f'C{r}'].font = accent_font; ws[f'C{r}'].alignment = Alignment(horizontal='center'); ws[f'C{r}'].border = thin
    ws[f'D{r}'] = row[2]; ws[f'D{r}'].font = normal_font; ws[f'D{r}'].alignment = Alignment(wrap_text=True); ws[f'D{r}'].border = thin
    ws.row_dimensions[r].height = 30
    r += 1
r += 1
# ── DRAFT PICK METHODOLOGY ──
ws.merge_cells(f'B{r}:D{r}')
ws[f'B{r}'] = 'DRAFT PICK VALUATION METHODOLOGY'
ws[f'B{r}'].font = section_font
ws[f'B{r}'].fill = gold_fill
r += 1

pick_explain = [
    'Draft pick values are NOT static numbers — they are calculated from YOUR leagues actual draft history.',
    '',
    'For each draft slot (1-112), the engine looks at every player drafted at that slot across 4 seasons and asks:',
    '  1. HIT RATE — What % of picks at this slot became top-15% players at their position?',
    '  2. STARTER RATE — What % became starter-level producers (above the starter line)?',
    '  3. AVG NORMALIZED PPG — Average production relative to the position starter threshold.',
    '',
    'These three metrics are combined into a single DHQ value for each pick slot.',
    '',
    'Key insight: An early 1st (slot 1-4) is worth 9,000-9,975 DHQ in your league.',
    'A late 1st (slot 13-16) is worth 5,688-6,600 — a massive difference.',
    'This is why "a 1st round pick" is not a single value — position within the round matters.',
    '',
    'War Room uses flat values (1st = 5,500 for all). DHQ uses slot-specific values derived from your leagues history.',
    'This means ReconAI and War Room will value the same pick differently until the engines are unified.',
]

for line in pick_explain:
    ws[f'B{r}'] = line
    ws[f'B{r}'].font = normal_font if line else small_font
    ws.merge_cells(f'B{r}:D{r}')
    r += 1
r += 1
# ── VALUE SCALE REFERENCE ──
ws.merge_cells(f'B{r}:D{r}')
ws[f'B{r}'] = 'VALUE SCALE REFERENCE'
ws[f'B{r}'].font = section_font
ws[f'B{r}'].fill = gold_fill
r += 1

scale_data = [
    ['DHQ Range', 'Tier', 'Examples'],
    ['8,000 - 10,000', 'ELITE — Franchise cornerstone', 'Josh Allen (9,350), Bijan Robinson (7,673)'],
    ['6,000 - 8,000', 'STAR — Top 20 dynasty asset', 'Ja\'Marr Chase (7,065), Trevor Lawrence (6,528)'],
    ['4,000 - 6,000', 'STARTER — Reliable weekly starter', 'CeeDee Lamb range, emerging young players'],
    ['2,000 - 4,000', 'DEPTH — Flex play / upside stash', 'Veteran starters, young backups with upside'],
    ['500 - 2,000', 'ROSTER FILLER — Bench depth', 'Backup QBs, aging veterans, IDP starters'],
    ['0 - 500', 'WAIVER LEVEL — Minimal value', 'Practice squad, late-round IDP, kickers'],
]

for i, row in enumerate(scale_data):
    for j, val in enumerate(row):
        c = ws.cell(row=r, column=j+2, value=val)
        c.border = thin
        if i == 0:
            c.font = header_font; c.fill = header_fill
        else:
            c.font = bold_font if j == 0 else normal_font
            c.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[r].height = 25
    r += 1
r += 1

# ── COLUMN DEFINITIONS ──
ws.merge_cells(f'B{r}:D{r}')
ws[f'B{r}'] = 'COLUMN DEFINITIONS (All Players Sheet)'
ws[f'B{r}'].font = section_font
ws[f'B{r}'].fill = gold_fill
r += 1
col_defs = [
    ['Column', 'Type', 'Definition'],
    ['Rank', 'Number', 'Overall DHQ rank among all 1,977 scored players'],
    ['Player', 'Text', 'Player full name'],
    ['Pos', 'Text', 'Position: QB, RB, WR, TE, DL, LB, DB, K'],
    ['Team', 'Text', 'Current NFL team (FA = free agent)'],
    ['Age', 'Number', 'Current age'],
    ['DHQ Value', 'Number', 'Final dynasty value on 0-10,000 scale'],
    ['wPPG', 'Decimal', 'Weighted points per game — production input to the formula'],
    ['AgeFactor', 'Decimal', '1.0 = in peak. <1.0 = declining. Decays 6%/year after peak age.'],
    ['SitMult', 'Decimal', 'Situation multiplier. >1.0 = elite rank. <1.0 = below average rank.'],
    ['Peak Yrs Left', 'Number', 'Years remaining in position-specific peak window'],
    ['Starter Seasons', 'Number', 'Number of seasons player produced starter-level output'],
    ['Recent GP', 'Number', 'Games played in most recent season (durability indicator)'],
    ['Trend %', 'Number', 'Year-over-year PPG change. +15 = 15% improvement. -20 = 20% decline.'],
]

for i, row in enumerate(col_defs):
    for j, val in enumerate(row):
        c = ws.cell(row=r, column=j+2, value=val)
        c.border = thin
        if i == 0:
            c.font = header_font; c.fill = header_fill
        else:
            c.font = bold_font if j == 0 else normal_font
            c.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[r].height = 22
    r += 1

# ── Save ──
outpath = '/Users/jacobc/Projects/reconai/reports/DHQ_Engine_Report.xlsx'
wb.save(outpath)
print(f'Saved with Formula Reference to {outpath} — {r} rows total on reference sheet')
