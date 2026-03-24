import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()

# Styles
gold = Font(name='Arial', bold=True, color='D4AF37', size=14)
header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='1A1A1A', end_color='1A1A1A', fill_type='solid')
data_font = Font(name='Arial', size=10)
bold_font = Font(name='Arial', bold=True, size=11)
normal_font = Font(name='Arial', size=11)
accent_font = Font(name='Arial', bold=True, size=11, color='2E86C1')
small_font = Font(name='Arial', size=10, color='666666')
pos_fills = {
    'QB': PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid'),
    'RB': PatternFill(start_color='E6FFE6', end_color='E6FFE6', fill_type='solid'),
    'WR': PatternFill(start_color='E6E6FF', end_color='E6E6FF', fill_type='solid'),
    'TE': PatternFill(start_color='FFF0E6', end_color='FFF0E6', fill_type='solid'),
    'DL': PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid'),
    'LB': PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid'),
    'DB': PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid'),
    'K':  PatternFill(start_color='FFFDE6', end_color='FFFDE6', fill_type='solid'),
}
thin = Border(left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
              top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'))
gold_fill = PatternFill(start_color='FFF8E1', end_color='FFF8E1', fill_type='solid')
gray_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
# Read data
with open('/Users/jacobc/Downloads/dhq_data_blended.txt', 'r') as f:
    content = f.read()
parts = content.split('=== PICKS ===')
player_lines = parts[0].replace('=== PLAYERS ===\n', '').strip().split('\n')
pick_lines = parts[1].strip().split('\n')
print(f"Players: {len(player_lines)-1}, Picks: {len(pick_lines)-1}")

# ═══ SHEET 1: All Players ═══
ws = wb.active
ws.title = 'All Players'
ws.merge_cells('A1:M1')
ws['A1'] = 'DHQ Engine Report — The Psycho League: Year VI (16-team SF IDP Half-PPR) — Blended Values'
ws['A1'].font = gold
ws['A1'].alignment = Alignment(horizontal='center')
ws.row_dimensions[1].height = 30

headers = ['Rank','Player','Pos','Team','Age','DHQ Value','wPPG','Age Factor','Sit Mult','Peak Yrs Left','Starter Seasons','Recent GP','Trend %']
for col, h in enumerate(headers, 1):
    c = ws.cell(row=3, column=col, value=h)
    c.font = header_font; c.fill = header_fill; c.alignment = Alignment(horizontal='center'); c.border = thin

for i, line in enumerate(player_lines[1:], 4):
    parts_p = line.split('|')
    if len(parts_p) < 13: continue
    vals = [int(parts_p[0]), parts_p[1], parts_p[2], parts_p[3], int(parts_p[4]),
            int(parts_p[5]), float(parts_p[6]), float(parts_p[7]), float(parts_p[8]),
            int(parts_p[9]), int(parts_p[10]), int(parts_p[11]), int(parts_p[12])]
    for col, v in enumerate(vals, 1):
        c = ws.cell(row=i, column=col, value=v)
        c.font = data_font; c.border = thin
        if col == 3:
            fill = pos_fills.get(v)
            if fill:
                for cc in range(1, 14): ws.cell(row=i, column=cc).fill = fill
widths = [6, 25, 5, 5, 5, 10, 7, 10, 9, 12, 14, 10, 8]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[chr(64+i) if i<=26 else 'A'+chr(64+i-26)].width = w
ws.freeze_panes = 'A4'
for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=6, max_col=6):
    for cell in row: cell.number_format = '#,##0'
for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=7, max_col=7):
    for cell in row: cell.number_format = '0.0'
for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=8, max_col=9):
    for cell in row: cell.number_format = '0.000'
print(f"Players sheet: {ws.max_row - 3} rows")

# ═══ SHEET 2: Draft Pick Values (Blended) ═══
ws2 = wb.create_sheet('Draft Pick Values')
ws2.merge_cells('A1:L1')
ws2['A1'] = 'DHQ Blended Pick Values — 60% League / 40% Industry (6 seasons of data)'
ws2['A1'].font = gold; ws2['A1'].alignment = Alignment(horizontal='center')
ws2.row_dimensions[1].height = 30

pick_headers = ['Slot','Round','Pick','Blended DHQ','League Raw','Industry','Lg Wt%','Ind Wt%','Hit Rate%','Starter%','Avg Norm PPG','Samples']
for col, h in enumerate(pick_headers, 1):
    c = ws2.cell(row=3, column=col, value=h)
    c.font = header_font; c.fill = header_fill; c.alignment = Alignment(horizontal='center'); c.border = thin

round_fills = {
    1: PatternFill(start_color='FFF8E1', end_color='FFF8E1', fill_type='solid'),
    2: PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid'),
    3: PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid'),
    4: PatternFill(start_color='F3E5F5', end_color='F3E5F5', fill_type='solid'),
    5: PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid'),
    6: PatternFill(start_color='ECEFF1', end_color='ECEFF1', fill_type='solid'),
    7: PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid'),
}
for i, line in enumerate(pick_lines[1:], 4):
    pp = line.split('|')
    if len(pp) < 12: continue
    vals = [int(pp[0]),int(pp[1]),int(pp[2]),int(pp[3]),int(pp[4]),int(pp[5]),
            int(pp[6]),int(pp[7]),int(pp[8]),int(pp[9]),float(pp[10]),int(pp[11])]
    rd = vals[1]
    for col, v in enumerate(vals, 1):
        c = ws2.cell(row=i, column=col, value=v)
        c.font = data_font; c.border = thin
        if rd in round_fills: c.fill = round_fills[rd]
for col, w in zip(range(1,13), [6,7,5,12,12,12,8,8,9,9,12,8]):
    ws2.column_dimensions[chr(64+col)].width = w
ws2.freeze_panes = 'A4'
for col in [4,5,6]:
    for row in ws2.iter_rows(min_row=4, max_row=ws2.max_row, min_col=col, max_col=col):
        for cell in row: cell.number_format = '#,##0'
print(f"Picks sheet: {ws2.max_row - 3} rows")
# ═══ SHEET 3: Formula Reference ═══
ws3 = wb.create_sheet('Formula Reference')
ws3.column_dimensions['A'].width = 4
ws3.column_dimensions['B'].width = 28
ws3.column_dimensions['C'].width = 14
ws3.column_dimensions['D'].width = 70
r = 1
ws3.merge_cells('B1:D1')
ws3['B1'] = 'DHQ ENGINE — FORMULA REFERENCE'
ws3['B1'].font = Font(name='Arial', bold=True, color='D4AF37', size=16)
ws3.row_dimensions[1].height = 35
r = 3

sections = [
    ('OVERVIEW', 'section', None),
    ('DHQ values every player on a 0-10000 scale using YOUR leagues scoring settings.', 'text', None),
    ('League: The Psycho League Year VI | 16-team | Superflex | IDP | Half-PPR', 'text', None),
    ('6 seasons of history | 448 draft picks | 862 FAAB transactions | 1977 players scored', 'text', None),
    ('', 'blank', None),
    ('MASTER FORMULA', 'section', None),
    ('DHQ = Core Score + Scarcity + Peak Bonus + Consistency + Durability', 'formula', None),
    ('Core Score = (wPPG x AgeFactor x SitMult) / TopComposite x 7500', 'subformula', None),
    ('', 'blank', None),
    ('COMPONENT BREAKDOWN', 'section', None),
]
for item in sections:
    text, style, _ = item
    if style == 'blank': r += 1; continue
    ws3.merge_cells(f'B{r}:D{r}')
    ws3[f'B{r}'] = text
    if style == 'section':
        ws3[f'B{r}'].font = Font(name='Arial', bold=True, color='D4AF37', size=13)
        ws3[f'B{r}'].fill = gold_fill
    elif style == 'formula':
        ws3[f'B{r}'].font = Font(name='Courier New', bold=True, size=13)
        ws3[f'B{r}'].fill = gray_fill
    elif style == 'subformula':
        ws3[f'B{r}'].font = Font(name='Courier New', size=11, color='444444')
    else:
        ws3[f'B{r}'].font = normal_font
    r += 1
# Component table
r += 1
for col, val in [('B','Component'), ('C','Weight'), ('D','How It Works')]:
    c = ws3[f'{col}{r}']; c.value = val; c.font = header_font; c.fill = header_fill; c.border = thin
r += 1
components = [
    ['CORE SCORE', '75%', 'Combines production (wPPG), age (AgeFactor), situation (SitMult). Normalized to 0-7500.'],
    ['  wPPG', '—', 'Weighted PPG across 5 seasons. Recent years weighted more. Best season gets bonus.'],
    ['  AgeFactor', '—', '1.0 in peak. Decays 6%/yr after peak. QB=34, RB=27, WR=30, TE=30.'],
    ['  SitMult', '—', 'Top3=1.20x, Top5=1.12x, Top10=1.05x, Bot25%=0.88x. Clamped 0.40-1.60.'],
    ['SCARCITY', '10%', 'Position premium. SF QB tiered: Top12=750, QB13-24=400, QB25+=100. 0 for unrostered.'],
    ['PEAK BONUS', '5%', '120 per peak year remaining, capped at 1000.'],
    ['CONSISTENCY', '~4%', '4+ starter seasons=400, 3=300, 2=150. 0 for unrostered.'],
    ['DURABILITY', '~1%', '16+ GP=100, 13+ GP=50.'],
]
for comp in components:
    is_sub = comp[0].startswith('  ')
    ws3[f'B{r}'] = comp[0]; ws3[f'B{r}'].font = normal_font if is_sub else bold_font; ws3[f'B{r}'].border = thin
    ws3[f'C{r}'] = comp[1]; ws3[f'C{r}'].font = accent_font; ws3[f'C{r}'].alignment = Alignment(horizontal='center'); ws3[f'C{r}'].border = thin
    ws3[f'D{r}'] = comp[2]; ws3[f'D{r}'].font = normal_font; ws3[f'D{r}'].alignment = Alignment(wrap_text=True); ws3[f'D{r}'].border = thin
    if not is_sub: ws3[f'B{r}'].fill = gray_fill; ws3[f'C{r}'].fill = gray_fill; ws3[f'D{r}'].fill = gray_fill
    ws3.row_dimensions[r].height = 30
    r += 1
# Blended Pick section
r += 1
ws3.merge_cells(f'B{r}:D{r}')
ws3[f'B{r}'] = 'BLENDED DRAFT PICK VALUES'
ws3[f'B{r}'].font = Font(name='Arial', bold=True, color='D4AF37', size=13)
ws3[f'B{r}'].fill = gold_fill
r += 1
blend_text = [
    'Pick values blend league-specific data with industry consensus (KTC, theScore, FantasyCalc, DLF).',
    'The weight auto-adjusts by league age — young leagues lean on industry, mature leagues trust their own data.',
    '',
    'CURRENT BLEND: 60% League / 40% Industry (6 seasons of history)',
    '',
    'Auto-Adjustment Schedule:',
    '  1-3 seasons:  80% industry / 20% league (small sample, trust the market)',
    '  4-5 seasons:  60% industry / 40% league (building confidence)',
    '  6-8 seasons:  40% industry / 60% league (YOUR LEAGUE NOW)',
    '  9+ seasons:   20% industry / 80% league (league data is reliable)',
    '',
    'This means as The Psycho League ages, pick values will increasingly reflect YOUR leagues',
    'actual draft outcomes rather than generic industry consensus. No code changes needed.',
    '',
    'Industry baseline derived from KTC crowdsourced data, theScore/Justin Boone rankings,',
    'FantasyCalc market values, and DLF dynasty trade analyzer — March 2026.',
]
for line in blend_text:
    ws3.merge_cells(f'B{r}:D{r}')
    ws3[f'B{r}'] = line
    ws3[f'B{r}'].font = bold_font if 'CURRENT' in line or 'Auto-Adjustment' in line else normal_font
    if 'CURRENT' in line: ws3[f'B{r}'].fill = gold_fill
    r += 1
# Value scale and worked example
r += 1
ws3.merge_cells(f'B{r}:D{r}')
ws3[f'B{r}'] = 'VALUE SCALE REFERENCE'
ws3[f'B{r}'].font = Font(name='Arial', bold=True, color='D4AF37', size=13)
ws3[f'B{r}'].fill = gold_fill
r += 1
for col, val in [('B','DHQ Range'), ('C','Tier'), ('D','Examples')]:
    c = ws3[f'{col}{r}']; c.value = val; c.font = header_font; c.fill = header_fill; c.border = thin
r += 1
scale = [
    ['8,000 - 10,000', 'ELITE', 'Josh Allen (9,350), Bijan Robinson (7,673)'],
    ['6,000 - 8,000', 'STAR', "Ja'Marr Chase (7,065), Trevor Lawrence (6,528)"],
    ['4,000 - 6,000', 'STARTER', 'CeeDee Lamb range, emerging young players'],
    ['2,000 - 4,000', 'DEPTH', 'Veteran starters, young backups with upside'],
    ['500 - 2,000', 'FILLER', 'Backup QBs, aging veterans, IDP starters'],
    ['0 - 500', 'WAIVER', 'Practice squad, late-round IDP, kickers'],
]
for row in scale:
    for j, val in enumerate(row):
        c = ws3.cell(row=r, column=j+2, value=val)
        c.border = thin; c.font = bold_font if j==0 else normal_font
    r += 1

# Save
outpath = '/Users/jacobc/Desktop/DHQ_Engine_Report_Blended.xlsx'
wb.save(outpath)
print(f'Saved to {outpath}')